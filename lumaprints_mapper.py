"""
Lumaprints Bulk Mapping Tool
Automates the process of mapping images to Lumaprints products
"""

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image
import os
from typing import List, Dict, Tuple
from data_manager_v3 import get_all_images

# Product templates for different aspect ratios
TEMPLATES = {
    "1:1": {
        "canvas": {
            "subcategory": "0.75in Stretched Canvas",
            "sizes": [(8, 8), (10, 10), (12, 12), (14, 14)],
            "options": [
                ("Canvas Border", "Mirror Wrap"),
                ("Canvas Hanging Hardware", "Sawtooth Hanger installed"),
                ("Canvas Finish", "Semi-Glossy")
            ]
        },
        "art_paper": {
            "subcategories": ["Hot Press Fine Art Paper", "Semi-Glossy Fine Art Paper", "Glossy Fine Art Paper"],
            "sizes": [(8, 8), (10, 10), (12, 12), (14, 14)],
            "options": [
                ("Bleed Size", "0.25in Bleed (0.25in on each side)")
            ]
        }
    },
    "3:2": {
        "canvas": {
            "subcategory": "0.75in Stretched Canvas",
            "sizes": [(8, 12), (12, 18), (16, 24), (20, 30), (24, 36)],
            "options": [
                ("Canvas Border", "Mirror Wrap"),
                ("Canvas Hanging Hardware", "Sawtooth Hanger installed"),
                ("Canvas Finish", "Semi-Glossy")
            ]
        },
        "art_paper": {
            "subcategories": ["Hot Press Fine Art Paper", "Semi-Glossy Fine Art Paper", "Glossy Fine Art Paper"],
            "sizes": [(8, 12), (12, 18), (16, 24), (20, 30), (24, 36)],
            "options": [
                ("Bleed Size", "0.25in Bleed (0.25in on each side)")
            ]
        }
    }
}


def detect_aspect_ratio(width: int, height: int) -> str:
    """
    Detect aspect ratio from image dimensions
    Returns: "1:1", "3:2", "2:3", "4:3", "3:4", "16:9", "9:16", or "other"
    """
    ratio = width / height
    
    # Allow 3% margin
    margin = 0.03
    
    if abs(ratio - 1.0) <= margin:
        return "1:1"
    elif abs(ratio - 1.5) <= margin:  # 3:2
        return "3:2"
    elif abs(ratio - 0.667) <= margin:  # 2:3
        return "2:3"
    elif abs(ratio - 1.333) <= margin:  # 4:3
        return "4:3"
    elif abs(ratio - 0.75) <= margin:  # 3:4
        return "3:4"
    elif abs(ratio - 1.778) <= margin:  # 16:9
        return "16:9"
    elif abs(ratio - 0.5625) <= margin:  # 9:16
        return "9:16"
    else:
        return "other"


def load_excel(filepath: str) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    """Load Excel workbook and return workbook and active sheet"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    return wb, ws


def sort_by_column_a(ws) -> None:
    """
    Sort worksheet by Column A (Product Name) alphabetically
    Keeps header row (row 1) in place
    """
    # Get all data rows (skip header)
    data_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row, col).value)
        data_rows.append(row_data)
    
    # Sort by first column (Column A)
    data_rows.sort(key=lambda x: x[0] if x[0] else "")
    
    # Write sorted data back
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row_idx, col_idx).value = value


def parse_size_from_option(option_text: str) -> Tuple[int, int]:
    """
    Parse size from option text like "Size - 8x8" or "Size - 12x18"
    Returns (width, length) tuple
    """
    import re
    
    if not option_text:
        return (0, 0)
    
    # Look for pattern like "8x8" or "12x18"
    match = re.search(r'(\d+)\s*[x×]\s*(\d+)', str(option_text))
    if match:
        return (int(match.group(1)), int(match.group(2)))
    
    return (0, 0)


def get_unmapped_products(ws) -> List[Dict]:
    """
    Get all unmapped products from worksheet
    Returns list of dicts with product info
    """
    unmapped = []
    
    for row in range(2, ws.max_row + 1):
        # Column O (15) = Mapping Status
        mapping_status = ws.cell(row, 15).value
        
        if mapping_status == "Unmapped":
            option2 = ws.cell(row, 3).value  # Column C
            width, length = parse_size_from_option(option2)
            
            product = {
                "row": row,
                "product_name": ws.cell(row, 1).value,  # Column A
                "option1": ws.cell(row, 2).value,  # Column B
                "option2": option2,  # Column C
                "size": f"{width}x{length}" if width > 0 else "Unknown",
                "width": width,
                "length": length,
                "sku": ws.cell(row, 10).value,  # Column J
                "existing_filename": ws.cell(row, 16).value,  # Column P
                "mapping_status": mapping_status
            }
            unmapped.append(product)
    
    return unmapped


def apply_mapping(ws, row: int, mapping_data: Dict) -> None:
    """
    Apply mapping data to a specific row
    
    mapping_data should contain:
    - product_handling: "Update", "Ignore", or "Custom"
    - image_filename: filename with extension
    - subcategory: product subcategory
    - width: product width in inches
    - length: product length in inches
    - options: list of (name, value) tuples
    """
    # Column S (19) - Product Handling
    ws.cell(row, 19).value = mapping_data.get("product_handling", "Update")
    
    # Column T (20) - Image Filename
    ws.cell(row, 20).value = mapping_data.get("image_filename", "")
    
    # Column U (21) - Subcategory
    ws.cell(row, 21).value = mapping_data.get("subcategory", "")
    
    # Column V (22) - Width
    ws.cell(row, 22).value = mapping_data.get("width", "")
    
    # Column W (23) - Length
    ws.cell(row, 23).value = mapping_data.get("length", "")
    
    # Columns X-AO (24-41) - Product Options
    options = mapping_data.get("options", [])
    for i, (option_name, option_value) in enumerate(options):
        if i < 9:  # Max 9 option pairs
            name_col = 24 + (i * 2)  # X, Z, AB, AD, AF, AH, AJ, AL, AN
            value_col = 25 + (i * 2)  # Y, AA, AC, AE, AG, AI, AK, AM, AO
            ws.cell(row, name_col).value = option_name
            ws.cell(row, value_col).value = option_value


def get_image_aspect_ratio(filename: str, images_dir: str = "/home/ubuntu/fifth-element-photography/data/images") -> str:
    """
    Get aspect ratio of an image file
    """
    filepath = os.path.join(images_dir, filename)
    
    if not os.path.exists(filepath):
        return "unknown"
    
    try:
        with Image.open(filepath) as img:
            width, height = img.size
            return detect_aspect_ratio(width, height)
    except Exception as e:
        print(f"Error reading image {filename}: {e}")
        return "unknown"


def get_available_images() -> List[Dict]:
    """
    Get list of available images from the gallery
    Returns list with filename, dimensions, and aspect ratio
    """
    images = get_all_images()
    
    available = []
    for img in images:
        filename = img.get("filename", "")
        dimensions = img.get("dimensions", "")
        
        # Parse dimensions
        if dimensions and "x" in dimensions:
            try:
                width, height = map(int, dimensions.split("x"))
                aspect_ratio = detect_aspect_ratio(width, height)
            except:
                aspect_ratio = "unknown"
        else:
            aspect_ratio = "unknown"
        
        available.append({
            "filename": filename,
            "title": img.get("title", filename),
            "dimensions": dimensions,
            "aspect_ratio": aspect_ratio
        })
    
    return available


def generate_mapping_template(aspect_ratio: str, product_type: str, size: Tuple[int, int]) -> Dict:
    """
    Generate mapping template based on aspect ratio and product type
    
    Args:
        aspect_ratio: "1:1" or "3:2"
        product_type: "canvas" or "art_paper"
        size: (width, length) tuple
    
    Returns:
        Dict with mapping data
    """
    if aspect_ratio not in TEMPLATES:
        return {}
    
    template = TEMPLATES[aspect_ratio].get(product_type, {})
    
    if product_type == "art_paper":
        # For art paper, use first subcategory as default
        subcategory = template.get("subcategories", [])[0] if template.get("subcategories") else ""
    else:
        subcategory = template.get("subcategory", "")
    
    return {
        "product_handling": "Update",
        "subcategory": subcategory,
        "width": size[0],
        "length": size[1],
        "options": template.get("options", [])
    }


def save_excel(wb: openpyxl.Workbook, output_path: str) -> None:
    """Save workbook to file"""
    wb.save(output_path)
