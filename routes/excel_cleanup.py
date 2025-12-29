"""
Excel Cleanup Tool
Standalone tool to sort and remove mapped rows from Lumaprints Excel exports
"""

from flask import Blueprint, request, jsonify, send_file
import openpyxl
import os

excel_cleanup_bp = Blueprint('excel_cleanup', __name__)

@excel_cleanup_bp.route('/api/excel-cleanup/process', methods=['POST'])
def process_excel():
    """Process Excel file: sort by Column A and remove mapped rows"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'File must be .xlsx format'}), 400
        
        # Save uploaded file temporarily
        temp_path = os.path.join('/tmp', 'excel_cleanup_upload.xlsx')
        file.save(temp_path)
        
        # Load workbook
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active
        
        # Get total rows before processing
        total_rows_before = ws.max_row - 1  # Exclude header
        
        # Step 1: Sort by Column A (Product Name)
        # Get all data rows (skip header row 1)
        data_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append({
                    'value': cell.value,
                    'style': cell._style
                })
            data_rows.append(row_data)
        
        # Sort by Column A (index 0)
        data_rows.sort(key=lambda x: str(x[0]['value']).lower() if x[0]['value'] else '')
        
        # Clear existing data (keep header)
        for row_idx in range(ws.max_row, 1, -1):
            ws.delete_rows(row_idx)
        
        # Write sorted data back
        for row_data in data_rows:
            new_row_idx = ws.max_row + 1
            for col_idx, cell_data in enumerate(row_data, start=1):
                cell = ws.cell(row=new_row_idx, column=col_idx)
                cell.value = cell_data['value']
                cell._style = cell_data['style']
        
        # Step 2: Delete rows with "Mapped" in Column O (column 15)
        deleted_count = 0
        row_idx = 2  # Start from first data row
        while row_idx <= ws.max_row:
            mapping_status = ws.cell(row=row_idx, column=15).value
            if mapping_status and str(mapping_status).strip().lower() == 'mapped':
                ws.delete_rows(row_idx)
                deleted_count += 1
            else:
                row_idx += 1
        
        # Get total rows after processing
        total_rows_after = ws.max_row - 1  # Exclude header
        
        # Save processed file
        output_path = os.path.join('/tmp', 'excel_cleanup_output.xlsx')
        wb.save(output_path)
        wb.close()
        
        return jsonify({
            'success': True,
            'total_rows_before': total_rows_before,
            'deleted_count': deleted_count,
            'total_rows_after': total_rows_after
        })
    
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel cleanup error: {error_details}")
        return jsonify({'error': str(e)}), 500

@excel_cleanup_bp.route('/api/excel-cleanup/download')
def download_cleaned_excel():
    """Download the cleaned Excel file"""
    try:
        output_path = os.path.join('/tmp', 'excel_cleanup_output.xlsx')
        if not os.path.exists(output_path):
            return jsonify({'error': 'No processed file found. Please process a file first.'}), 404
        
        return send_file(
            output_path,
            as_attachment=True,
            download_name='cleaned_lumaprints_export.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
