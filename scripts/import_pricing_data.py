#!/usr/bin/env python3.11
"""
Fifth Element Photography - Pricing Data Import Script
Imports product pricing from Excel spreadsheet into SQLite database
Version: Beta 0.1.0
Date: October 28, 2025
"""

import sqlite3
import openpyxl
import sys
import os
from decimal import Decimal
from datetime import datetime

# Database path
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'print_ordering.db')
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'LumaprintsCatalogandSizingw_aspectratios.xlsx')

# Product mapping: Excel column name → subcategory_id
CANVAS_MAPPING = {
    '1.25IN STRETCHED CANVAS': 101002,
    '1.5IN STRETCHED CANVAS': 101003,
    '0.75IN STRETCHED CANVAS': 101001,
    'ROLLED CANVAS': 101005
}

FRAMED_CANVAS_MAPPING = {
    '0.75IN FRAMED CANVAS': 102001,
    '1.25IN FRAMED CANVAS': 102002,
    '1.50IN FRAMED CANVAS': 102003
}

FINE_ART_PAPER_MAPPING = {
    'HOT PRESS': 103002,
    'COLD PRESS': 103003,
    'SEMI-GLOSS': 103005,
    'GLOSSY': 103007
}

# Foam-mounted products - need to verify these subcategory IDs with Lumaprints API
FOAM_MOUNTED_MAPPING = {
    'HOT PRESS': 104002,  # Placeholder - verify with API
    'COLD PRESS': 104003,  # Placeholder - verify with API
    'SEMI-GLOSS': 104005,  # Placeholder - verify with API
    'GLOSSY': 104007       # Placeholder - verify with API
}

def init_database():
    """Initialize database with schema"""
    print("Initializing database...")
    
    # Read schema file
    schema_path = os.path.join(os.path.dirname(__file__), '..', 'database', 'print_ordering_schema.sql')
    with open(schema_path, 'r') as f:
        schema_sql = f.read()
    
    # Create database and execute schema
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.executescript(schema_sql)
    conn.commit()
    
    print(f"✓ Database initialized at {DB_PATH}")
    return conn

def parse_size(size_str):
    """Parse size string like '12×18"' into width and height integers"""
    if not size_str:
        return None, None
    
    # Remove quotes and split by × or x
    size_str = size_str.replace('"', '').replace('″', '')
    parts = size_str.replace('×', 'x').split('x')
    
    if len(parts) != 2:
        return None, None
    
    try:
        width = int(parts[0].strip())
        height = int(parts[1].strip())
        return width, height
    except ValueError:
        return None, None

def get_or_create_aspect_ratio(cursor, width, height):
    """Get or create aspect ratio for given dimensions"""
    # Calculate aspect ratio
    ratio_decimal = width / height if height > width else height / width
    
    # Determine ratio name
    if abs(ratio_decimal - 1.0) < 0.01:
        ratio_name = '1:1'
    elif abs(ratio_decimal - 1.5) < 0.01:
        ratio_name = '3:2'
    elif abs(ratio_decimal - 1.333) < 0.01:
        ratio_name = '4:3'
    elif abs(ratio_decimal - 1.25) < 0.01:
        ratio_name = '5:4'
    elif abs(ratio_decimal - 1.778) < 0.01:
        ratio_name = '16:9'
    else:
        ratio_name = f'{width}:{height}'
    
    # Check if aspect ratio exists
    cursor.execute('SELECT aspect_ratio_id FROM aspect_ratios WHERE ratio_name = ?', (ratio_name,))
    row = cursor.fetchone()
    
    if row:
        return row[0]
    
    # Create new aspect ratio
    cursor.execute('''
        INSERT INTO aspect_ratios (ratio_name, ratio_decimal, display_name, description)
        VALUES (?, ?, ?, ?)
    ''', (ratio_name, ratio_decimal, ratio_name, f'{ratio_name} aspect ratio'))
    
    return cursor.lastrowid

def get_or_create_size(cursor, width, height, aspect_ratio_id):
    """Get or create print size"""
    size_name = f'{width}×{height}"'
    
    # Check if size exists
    cursor.execute('SELECT size_id FROM print_sizes WHERE width = ? AND height = ?', (width, height))
    row = cursor.fetchone()
    
    if row:
        return row[0]
    
    # Create new size
    cursor.execute('''
        INSERT INTO print_sizes (aspect_ratio_id, width, height, size_name)
        VALUES (?, ?, ?, ?)
    ''', (aspect_ratio_id, width, height, size_name))
    
    return cursor.lastrowid

def import_pricing_section(cursor, ws, start_row, product_mapping, aspect_ratio_name, header_row=None):
    """Import pricing for a section of the spreadsheet"""
    print(f"\nImporting {aspect_ratio_name} pricing...")
    
    # Find header row (contains product names) if not provided
    if header_row is None:
        for row_idx in range(start_row, start_row + 5):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value == 'SIZE':
                header_row = row_idx
                break
    
    if not header_row:
        print(f"  ⚠ Could not find header row starting from row {start_row}")
        return 0
    
    # Read column headers (product names)
    headers = []
    for col_idx in range(2, 20):  # Check up to column 20
        header_value = ws.cell(row=header_row, column=col_idx).value
        if header_value:
            headers.append((col_idx, header_value))
    
    print(f"  Found {len(headers)} products: {[h[1] for h in headers]}")
    
    # Import pricing data
    imported_count = 0
    row_idx = header_row + 1
    
    while True:
        size_cell = ws.cell(row=row_idx, column=1).value
        
        # Stop if we hit an empty row or new section
        if not size_cell or size_cell in ['1:1 Aspect Ratio', '3:2 Aspect Ratio', 'Framed Canvas', 'Fine Art Paper', 'Foam-mounted Fine Art Paper']:
            break
        
        # Parse size
        width, height = parse_size(size_cell)
        if not width or not height:
            row_idx += 1
            continue
        
        # Get or create aspect ratio and size
        aspect_ratio_id = get_or_create_aspect_ratio(cursor, width, height)
        size_id = get_or_create_size(cursor, width, height, aspect_ratio_id)
        
        # Import prices for each product
        for col_idx, product_name in headers:
            price_value = ws.cell(row=row_idx, column=col_idx).value
            
            # Get subcategory ID
            subcategory_id = product_mapping.get(product_name)
            if not subcategory_id:
                continue
            
            # Handle "n/a" or missing prices
            is_available = True
            cost_price = 0.0
            
            if price_value == 'n/a' or price_value is None:
                is_available = False
            else:
                try:
                    cost_price = float(price_value)
                except (ValueError, TypeError):
                    is_available = False
            
            # Insert or update pricing
            cursor.execute('''
                INSERT OR REPLACE INTO base_pricing 
                (subcategory_id, size_id, cost_price, is_available, notes)
                VALUES (?, ?, ?, ?, ?)
            ''', (subcategory_id, size_id, cost_price, is_available, 
                  'n/a in source data' if not is_available else None))
            
            imported_count += 1
        
        row_idx += 1
    
    print(f"  ✓ Imported {imported_count} pricing entries")
    return imported_count

def import_all_pricing(conn):
    """Import all pricing data from Excel"""
    print("\n" + "="*60)
    print("IMPORTING PRICING DATA FROM EXCEL")
    print("="*60)
    
    # Load workbook
    print(f"\nLoading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    print(f"✓ Loaded worksheet: {ws.title}")
    
    cursor = conn.cursor()
    total_imported = 0
    
    # Import Canvas - 1:1 Aspect Ratio (starts around row 3)
    total_imported += import_pricing_section(cursor, ws, 1, CANVAS_MAPPING, "Canvas 1:1")
    
    # Import Canvas - 3:2 Aspect Ratio (starts around row 10, uses same headers as 1:1)
    total_imported += import_pricing_section(cursor, ws, 10, CANVAS_MAPPING, "Canvas 3:2", header_row=3)
    
    # Import Framed Canvas - 1:1 Aspect Ratio (starts around row 20)
    total_imported += import_pricing_section(cursor, ws, 20, FRAMED_CANVAS_MAPPING, "Framed Canvas 1:1")
    
    # Import Framed Canvas - 3:2 Aspect Ratio (starts around row 30)
    total_imported += import_pricing_section(cursor, ws, 30, FRAMED_CANVAS_MAPPING, "Framed Canvas 3:2")
    
    # Import Fine Art Paper - 1:1 Aspect Ratio (starts around row 41)
    total_imported += import_pricing_section(cursor, ws, 41, FINE_ART_PAPER_MAPPING, "Fine Art Paper 1:1")
    
    # Import Fine Art Paper - 3:2 Aspect Ratio (starts around row 53)
    total_imported += import_pricing_section(cursor, ws, 53, FINE_ART_PAPER_MAPPING, "Fine Art Paper 3:2")
    
    # Import Foam-mounted - 1:1 Aspect Ratio (starts around row 63)
    total_imported += import_pricing_section(cursor, ws, 63, FOAM_MOUNTED_MAPPING, "Foam-mounted 1:1")
    
    # Import Foam-mounted - 3:2 Aspect Ratio (starts around row 75)
    total_imported += import_pricing_section(cursor, ws, 75, FOAM_MOUNTED_MAPPING, "Foam-mounted 3:2")
    
    conn.commit()
    
    print("\n" + "="*60)
    print(f"✓ IMPORT COMPLETE: {total_imported} total pricing entries")
    print("="*60)
    
    return total_imported

def set_default_markup(conn):
    """Set default global markup rule (DISABLED - user will set manually)"""
    print("\nNo default markup set - all prices at cost")
    print("Use admin interface to configure markup rules")

def print_summary(conn):
    """Print import summary"""
    cursor = conn.cursor()
    
    print("\n" + "="*60)
    print("DATABASE SUMMARY")
    print("="*60)
    
    # Count aspect ratios
    cursor.execute('SELECT COUNT(*) FROM aspect_ratios')
    print(f"\nAspect Ratios: {cursor.fetchone()[0]}")
    
    # Count sizes
    cursor.execute('SELECT COUNT(*) FROM print_sizes')
    print(f"Print Sizes: {cursor.fetchone()[0]}")
    
    # Count pricing entries
    cursor.execute('SELECT COUNT(*) FROM base_pricing WHERE is_available = TRUE')
    available_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM base_pricing WHERE is_available = FALSE')
    unavailable_count = cursor.fetchone()[0]
    print(f"Pricing Entries: {available_count} available, {unavailable_count} unavailable")
    
    # Count by category
    print("\nPricing by Category:")
    cursor.execute('''
        SELECT pc.display_name, COUNT(*) as count
        FROM base_pricing bp
        JOIN product_subcategories ps ON bp.subcategory_id = ps.subcategory_id
        JOIN product_categories pc ON ps.category_id = pc.category_id
        WHERE bp.is_available = TRUE
        GROUP BY pc.category_id
        ORDER BY pc.display_order
    ''')
    for row in cursor.fetchall():
        print(f"  {row[0]}: {row[1]} entries")
    
    # Sample retail prices
    print("\nSample Retail Prices (with 40% markup):")
    cursor.execute('''
        SELECT ps.display_name, pz.size_name, bp.cost_price, rp.retail_price
        FROM retail_pricing rp
        JOIN base_pricing bp ON rp.pricing_id = bp.pricing_id
        JOIN product_subcategories ps ON bp.subcategory_id = ps.subcategory_id
        JOIN print_sizes pz ON bp.size_id = pz.size_id
        LIMIT 10
    ''')
    for row in cursor.fetchall():
        print(f"  {row[0]} {row[1]}: ${row[2]:.2f} cost → ${row[3]:.2f} retail")
    
    print("\n" + "="*60)

def main():
    """Main import process"""
    try:
        # Initialize database
        conn = init_database()
        
        # Import pricing data
        import_all_pricing(conn)
        
        # Note: No default markup - user configures via admin
        set_default_markup(conn)
        
        # Print summary
        print_summary(conn)
        
        conn.close()
        
        print("\n✅ Import completed successfully!")
        print(f"Database location: {DB_PATH}")
        
    except Exception as e:
        print(f"\n❌ Error during import: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()

